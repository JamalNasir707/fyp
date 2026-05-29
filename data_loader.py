import csv
import os
from typing import List, Dict, Any, Optional
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


def get_category_score(category: str) -> int:
    if not category:
        return 6

    category = category.lower().strip()

    if "mountain" in category:
        return 9
    if "valley" in category:
        return 9
    if "lake" in category:
        return 8
    if "waterfall" in category:
        return 8
    if "island" in category:
        return 8
    if "national" in category:
        return 8
    if "mosque" in category:
        return 7
    if "fort" in category:
        return 7
    if "hill" in category:
        return 7
    if "coastal" in category:
        return 7

    return 6


def get_visit_duration(category: str) -> float:
    """
    Estimated realistic visit duration (in hours)
    based on attraction type.
    """

    if not category:
        return 1.5

    category = category.lower().strip()

    if "mountain" in category:
        return 3.0
    if "valley" in category:
        return 2.5
    if "lake" in category:
        return 2.0
    if "waterfall" in category:
        return 1.5
    if "mosque" in category:
        return 1.0
    if "fort" in category:
        return 1.5
    if "hill" in category:
        return 2.0

    return 1.5


def get_district_cost(district: str) -> int:
    if not district:
        return 150

    district = district.lower().strip()

    high_cost_regions = [
        "gilgit", "skardu", "hunza", "astore", "khaplu"
    ]

    medium_cost_regions = [
        "khyber", "azad", "balochistan"
    ]

    for region in high_cost_regions:
        if region in district:
            return 500

    for region in medium_cost_regions:
        if region in district:
            return 300

    return 150


def _safe_float(v) -> Optional[float]:
    try:
        if v is None:
            return None
        return float(v)
    except Exception:
        return None


def _normalize_province(p: Any) -> str:
    s = "" if p is None else str(p).strip()
    s = s.replace("\u2010", "-")
    s = s.replace("\u2011", "-")
    s = s.replace("\u2012", "-")
    s = s.replace("\u2013", "-")
    s = s.replace("\u2014", "-")
    s = s.replace("\u2015", "-")
    s = s.replace("\u2212", "-")
    return s


def _load_locations_from_xlsx_v2_if_available() -> Optional[List[Dict[str, Any]]]:
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
    xlsx_path = os.path.join(base_dir, "data", "travel_intelligence_dataset_v2.xlsx")

    if not os.path.exists(xlsx_path):
        return None
    if load_workbook is None:
        return None

    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.rows)
    if not rows:
        return []

    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(headers)}

    required = ["latitude", "longitude", "category", "province", "popularity_score"]
    if not all(h in idx for h in required):
        return None

    locations: List[Dict[str, Any]] = []

    for r in rows[1:]:
        def val(key: str):
            i = idx.get(key)
            return r[i].value if i is not None and i < len(r) else None

        raw_name = val("name") if "name" in idx else None
        lat = _safe_float(val("latitude"))
        lon = _safe_float(val("longitude"))
        category = val("category") if val("category") is not None else ""
        province = _normalize_province(val("province"))
        city_val = val("city") if "city" in idx else None
        city = (str(city_val).strip() if city_val is not None else "")
        visit_duration = None
        if "average_visit_duration" in idx:
            visit_duration = _safe_float(val("average_visit_duration"))
        elif "visit_duration_hours" in idx:
            visit_duration = _safe_float(val("visit_duration_hours"))
        popularity = _safe_float(val("popularity_score"))
        safety = _safe_float(val("safety_rating")) if "safety_rating" in idx else None
        if "transport_access" in idx:
            transport = _safe_float(val("transport_access"))
        elif "transport_accessibility" in idx:
            transport = _safe_float(val("transport_accessibility"))
        else:
            transport = None
        avg_cost = _safe_float(val("average_cost")) if "average_cost" in idx else None

        if lat is None or lon is None:
            continue

        if raw_name is None or str(raw_name).strip() == "":
            synthesized = f"{str(category).strip() or 'Attraction'} - {city or province}"
            name = synthesized
        else:
            name = str(raw_name).strip()

        cat_score = get_category_score(str(category))
        final_score = (
            0.5 * (popularity if popularity is not None else 0.0)
            + 0.2 * (safety if safety is not None else 0.0)
            + 0.2 * (transport if transport is not None else 0.0)
            + 0.1 * cat_score
        )

        location = {
            "name": str(name).strip(),
            "lat": lat,
            "lon": lon,
            "category": str(category).strip(),
            "province": province,
            "city": city,
            "visit_duration": visit_duration if visit_duration is not None else get_visit_duration(str(category)),
            "score": final_score,
            "cost": avg_cost if avg_cost is not None else get_district_cost(province),
        }

        locations.append(location)

    return locations


def load_locations_from_csv(use_live_data: bool = False):
    locations = []
    v2 = _load_locations_from_xlsx_v2_if_available()
    if v2 is not None:
        locations.extend(v2)
    else:
        file_path = os.path.join(os.path.dirname(__file__), "tourist-destinations-in-pakistan.csv")
        with open(file_path, newline='', encoding="utf-8") as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                try:
                    district = row.get("district", "")
                    province = _normalize_province(district)
                    category = row.get("category", "").strip()

                    if not province or not row.get("latitude") or not row.get("longitude"):
                        continue

                    location = {
                        "name": row.get("_key", "").strip(),
                        "lat": float(row["latitude"]),
                        "lon": float(row["longitude"]),
                        "category": category,
                        "score": get_category_score(category),
                        "cost": get_district_cost(province),
                        "visit_duration": get_visit_duration(category),
                        "province": province,
                        "city": ""
                    }
                    locations.append(location)
                except Exception:
                    continue

    if use_live_data:
        dynamic_file_path = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir, "data", "places_dynamic.csv"))
        if os.path.exists(dynamic_file_path):
            with open(dynamic_file_path, newline='', encoding="utf-8") as csvfile:
                reader = csv.DictReader(csvfile)
                for row in reader:
                    try:
                        category = row.get("category", "").strip()
                        cat_score = get_category_score(category)
                        popularity = _safe_float(row.get("popularity_score")) or 5.0
                        safety = _safe_float(row.get("safety_rating")) or 8.0
                        
                        final_score = (
                            0.5 * popularity
                            + 0.2 * safety
                            + 0.2 * 5.0 # transport default
                            + 0.1 * cat_score
                        )
                        
                        location = {
                            "name": row.get("name", "").strip(),
                            "lat": float(row["latitude"]),
                            "lon": float(row["longitude"]),
                            "category": category,
                            "score": final_score,
                            "cost": _safe_float(row.get("average_cost")) or 2000,
                            "visit_duration": get_visit_duration(category),
                            "province": _normalize_province(row.get("province", "")),
                            "city": row.get("city", "").strip()
                        }
                        locations.append(location)
                    except Exception:
                        continue
    
    # Deduplicate by name just in case
    seen_names = set()
    unique_locations = []
    for loc in locations:
        n = loc["name"].lower()
        if n not in seen_names:
            seen_names.add(n)
            unique_locations.append(loc)

    return unique_locations

def load_master_dataset() -> List[Dict[str, Any]]:
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
    master_path = os.path.join(base_dir, "data", "pakistan_master_dataset.csv")
    locations: List[Dict[str, Any]] = []
    if not os.path.exists(master_path):
        return locations
    with open(master_path, newline='', encoding="utf-8") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            try:
                name = (row.get("name") or "").strip()
                lat = _safe_float(row.get("latitude"))
                lon = _safe_float(row.get("longitude"))
                category = (row.get("category") or "").strip()
                province = _normalize_province(row.get("province"))
                city = (row.get("city") or "").strip()
                avg_cost = _safe_float(row.get("average_cost"))
                rating = _safe_float(row.get("rating")) or 0.0
                popularity = _safe_float(row.get("popularity_score")) or 0.0
                importance = int(float(row.get("importance_level") or 1))
                if not name or lat is None or lon is None or not province:
                    continue
                cost_val = avg_cost if avg_cost is not None else get_district_cost(province)
                location = {
                    "name": name,
                    "lat": lat,
                    "lon": lon,
                    "category": category,
                    "province": province,
                    "city": city,
                    "visit_duration": get_visit_duration(category),
                    "cost": cost_val,
                    "rating": rating,
                    "popularity_score": popularity,
                    "importance_level": importance,
                    "score": 0.0
                }
                locations.append(location)
            except Exception:
                continue
    seen = set()
    unique_locations = []
    for loc in locations:
        key = loc["name"].lower()
        if key not in seen:
            seen.add(key)
            unique_locations.append(loc)
    return unique_locations
